'''The pantry: seeded from the spreadsheet, read from the database.

``PantrySheet`` parses the two sheets and produces rows. ``PantryRepository``
stores them once and answers every later question from the database, so the
running system never depends on a file being mounted.
'''

from __future__ import annotations

import re
from collections.abc import Iterable
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from .database import DatabaseUnavailable
from .units import Dimension, Package, UnitConverter, UnknownUnitError


class SpreadsheetError(RuntimeError):
    '''Raised when the workbook is missing, unreadable or malformed.'''


@dataclass
class PantryItem:
    '''One ingredient, with both sheets already reconciled.'''

    name: str
    stock: float
    base_unit: str
    dimension: str
    unit_cost: float
    price_paid: float
    sheet_unit: str
    package_label: str | None
    priced_per_piece: bool
    # What committed batches already took out of it. Kept beside the stock so a
    # message can say "você tinha 1,5 kg e a lasanha levou 1 kg" instead of
    # announcing a number with no history.
    used: float = 0.0
    seeded_stock: float = 0.0
    key: str = field(init=False)

    def __post_init__(self) -> None:
        self.key = UnitConverter.normalise_text(self.name)

    def as_dict(self) -> dict:
        payload = {
            'ingredient': self.name,
            'stock': round(self.stock, 4),
            'already_committed': round(self.used, 4),
            'stock_before_any_dish': round(self.seeded_stock, 4),
            'unit': self.base_unit,
            'unit_cost': round(self.unit_cost, 4),
            'unit_cost_label': f'R$ {self.unit_cost:.2f}/{self.base_unit}',
            'price_paid': round(self.price_paid, 2),
            'sheet_unit': self.sheet_unit,
            'search_keyword': self.search_keyword,
        }
        if self.package_label:
            payload['package'] = self.package_label
        if self.priced_per_piece:
            payload['note'] = (
                'Priced per loose piece. Fine for recipes that count units; if a '
                'recipe asks for g/ml, ask Dona Maria the package weight and call '
                'record_package_size.'
            )
        return payload

    @property
    def search_keyword(self) -> str:
        '''Ingredient name trimmed down to what a cook would actually type.

        Drops the sheet parenthetical and grading noise ('tipo 1'), which only
        narrow a web search without adding meaning.
        '''
        trimmed = self.name.split('(')[0]
        trimmed = re.sub(r'\btipo\s+\d+\b', '', trimmed, flags=re.IGNORECASE)
        return re.sub(r'\s+', ' ', trimmed).strip()


class PantrySheet:
    '''Reads the spreadsheet. Its only job is to seed the database.

    The application does not read this file: it reads Postgres. Keeping the
    reader separate means the parsing can be tested without a database, and the
    running system never depends on a file being mounted.
    '''

    STOCK_SHEET = 'Despensa'
    PRICE_SHEET = 'Precos'

    def __init__(self, spreadsheet_path):
        self.spreadsheet_path = Path(spreadsheet_path)

    def rows(self) -> list[dict]:
        '''The two sheets joined, as raw values ready to be stored.'''
        if not self.spreadsheet_path.exists():
            raise SpreadsheetError(f'spreadsheet not found at {self.spreadsheet_path}')

        workbook = load_workbook(self.spreadsheet_path, data_only=True, read_only=True)
        try:
            missing = {self.STOCK_SHEET, self.PRICE_SHEET} - set(workbook.sheetnames)
            if missing:
                raise SpreadsheetError(
                    f'missing sheets {sorted(missing)}; found {workbook.sheetnames}'
                )
            stock_rows = self._read_rows(workbook[self.STOCK_SHEET])
            price_rows = self._read_rows(workbook[self.PRICE_SHEET])
        finally:
            workbook.close()

        stock_by_key = {}
        for row in stock_rows:
            name = self._column(row, 'ingrediente')
            if name:
                stock_by_key[UnitConverter.normalise_text(name)] = (
                    float(self._column(row, 'quantidade') or 0),
                    str(self._column(row, 'unidade') or ''),
                )

        joined = []
        for row in price_rows:
            name = self._column(row, 'ingrediente')
            if not name:
                continue
            key = UnitConverter.normalise_text(name)
            bought = float(self._column(row, 'quantidade') or 0)
            if bought <= 0:
                continue
            bought_unit = str(self._column(row, 'unidade') or '')
            stock_quantity, stock_unit = stock_by_key.get(key, (0.0, bought_unit))
            joined.append(
                {
                    'ingredient_key': key,
                    'ingredient': str(name).strip(),
                    'stock_quantity': stock_quantity,
                    'stock_unit': stock_unit or bought_unit,
                    'bought_quantity': bought,
                    'bought_unit': bought_unit,
                    'price_paid': float(self._column(row, 'preco') or 0),
                }
            )
        return joined

    @staticmethod
    def _read_rows(sheet) -> list[dict]:
        rows = sheet.iter_rows(values_only=True)
        try:
            header = [UnitConverter.normalise_text(cell or '') for cell in next(rows)]
        except StopIteration:
            return []
        return [
            dict(zip(header, row))
            for row in rows
            if any(cell is not None and str(cell).strip() for cell in row)
        ]

    @staticmethod
    def _column(row: dict, *prefixes: str):
        for key, value in row.items():
            if any(key.startswith(prefix) for prefix in prefixes):
                return value
        return None


class PantryRepository:
    """The pantry as the application sees it: rows in Postgres.

    The spreadsheet seeds this once. Nothing here reads a file, so the running
    system does not depend on one being mounted, and a package weight she tells
    us later is applied on top of what was seeded rather than editing it.
    """

    # Connectives carry no meaning when matching names. Counting them made
    # 'farinha de rosca' score 0.5 against 'farinha de trigo' and match.
    STOP_WORDS = frozenset({'de', 'da', 'do', 'em', 'com', 'e', 'a', 'o'})
    MATCH_THRESHOLD = 0.5

    def __init__(self, db, seed_from=None):
        self.db = db
        self.seed_from = Path(seed_from) if seed_from else None
        self.package_sizes: dict[str, dict] = {}
        self.items: dict[str, PantryItem] = {}
        self.reload()

    # ------------------------------------------------------------- seeding

    def seed(self, force: bool = False) -> dict:
        """Load the spreadsheet into Postgres. Idempotent unless forced."""
        if self.seed_from is None:
            return {'seeded': False, 'reason': 'no spreadsheet configured'}

        existing = self.db.one('SELECT count(*) AS n FROM pantry_items')
        if existing and int(existing['n']) and not force:
            return {'seeded': False, 'reason': 'already seeded', 'items': int(existing['n'])}

        rows = PantrySheet(self.seed_from).rows()
        with self.db.connect() as conn:
            if force:
                conn.execute('TRUNCATE pantry_items')
            conn.cursor().executemany(
                """INSERT INTO pantry_items
                       (ingredient_key, ingredient, stock_quantity, stock_unit,
                        bought_quantity, bought_unit, price_paid)
                    VALUES (%(ingredient_key)s, %(ingredient)s, %(stock_quantity)s,
                            %(stock_unit)s, %(bought_quantity)s, %(bought_unit)s,
                            %(price_paid)s)
                   ON CONFLICT (ingredient_key) DO UPDATE SET
                       ingredient = EXCLUDED.ingredient,
                       stock_quantity = EXCLUDED.stock_quantity,
                       stock_unit = EXCLUDED.stock_unit,
                       bought_quantity = EXCLUDED.bought_quantity,
                       bought_unit = EXCLUDED.bought_unit,
                       price_paid = EXCLUDED.price_paid,
                       seeded_at = now()""",
                rows,
            )
            conn.commit()
        self.reload()
        return {'seeded': True, 'items': len(rows), 'source': str(self.seed_from)}

    # ------------------------------------------------------------- loading

    def reload(self) -> None:
        """Rebuild every derived unit cost from what is stored."""
        self._load_package_sizes()
        try:
            rows = self.db.query('SELECT * FROM pantry_items ORDER BY ingredient')
        except DatabaseUnavailable:
            # Unreachable is empty, not wrong. Anything else is a bug and must
            # surface: a broad except here once turned a mapping error into a
            # silently empty pantry.
            self.items = {}
            return

        used = self._committed_usage()
        items: dict[str, PantryItem] = {}
        for row in rows:
            key = row['ingredient_key']
            bought = float(row['bought_quantity'])
            package = self._package_for(key, row['bought_unit'])
            if bought <= 0 or package.factor <= 0:
                continue
            stock_package = self._package_for(key, row['stock_unit'] or row['bought_unit'])
            price = float(row['price_paid'])

            seeded = float(row['stock_quantity']) * stock_package.factor
            spent = used.get(key, 0.0)

            items[key] = PantryItem(
                name=row['ingredient'],
                # What is left, never what was bought. A dish committed earlier
                # already took its share out, and the next dish has to see the
                # fridge as it is now.
                stock=max(seeded - spent, 0.0),
                used=min(spent, seeded),
                seeded_stock=seeded,
                base_unit=package.base_unit,
                dimension=package.dimension,
                unit_cost=price / (bought * package.factor),
                price_paid=price,
                sheet_unit=row['bought_unit'],
                package_label=row['bought_unit'] if package.is_packaged else None,
                priced_per_piece=(
                    package.dimension == Dimension.COUNT and key not in self.package_sizes
                ),
            )
        self.items = items

    def _committed_usage(self) -> dict[str, float]:
        """How much of each ingredient the dishes on the menu already ate."""
        try:
            rows = self.db.query(
                """SELECT ingredient_key, sum(quantity) AS spent
                     FROM pantry_usage
                 GROUP BY ingredient_key"""
            )
        except DatabaseUnavailable:
            return {}
        return {row['ingredient_key']: float(row['spent']) for row in rows}

    def record_usage(
        self, dish: str, lines: Iterable[tuple[str, float]], portions: int
    ) -> dict[str, object]:
        """Take a committed batch out of the pantry.

        Append-only on purpose: the seeded row is never rewritten, so the stock
        she started with stays readable next to what each dish consumed. Called
        once per dish, when the dish is accepted onto the menu - not when it is
        merely costed, because pricing a dish she never sells should not empty
        her fridge.
        """
        # One line per ingredient, not per recipe line. A recipe can ask for
        # tomato twice - the fresh one and the one that becomes the sauce - and
        # two rows would make the story read "levou 1 kg e levou 1,8 kg" instead
        # of "levou 2,8 kg".
        totals: dict[str, float] = {}
        for key, quantity in lines:
            if float(quantity) > 0:
                totals[key] = totals.get(key, 0.0) + float(quantity)

        rows = [
            {
                'dish': dish,
                'ingredient_key': key,
                'quantity': quantity,
                'base_unit': self.items[key].base_unit if key in self.items else '',
                'portions': int(portions),
            }
            for key, quantity in totals.items()
        ]
        if not rows:
            return {'recorded': False, 'reason': 'nothing to take out of the pantry'}

        with self.db.connect() as conn:
            # Accepting the same dish twice must not eat her pantry twice. The
            # dish owns its rows: writing them again replaces them.
            conn.cursor().execute(
                'DELETE FROM pantry_usage WHERE dish = %(dish)s', {'dish': dish}
            )
            conn.cursor().executemany(
                """INSERT INTO pantry_usage
                       (dish, ingredient_key, quantity, base_unit, portions)
                    VALUES (%(dish)s, %(ingredient_key)s, %(quantity)s,
                            %(base_unit)s, %(portions)s)""",
                rows,
            )
            conn.commit()
        self.reload()
        return {
            'recorded': True,
            'dish': dish,
            'ingredients': len(rows),
            'left': {
                row['ingredient_key']: round(self.items[row['ingredient_key']].stock, 4)
                for row in rows
                if row['ingredient_key'] in self.items
            },
        }

    def usage_history(self, key: str) -> list[dict[str, object]]:
        """Which dishes ate this ingredient, oldest first."""
        try:
            rows = self.db.query(
                """SELECT dish, quantity, base_unit, portions, committed_at
                     FROM pantry_usage
                    WHERE ingredient_key = %(key)s
                 ORDER BY id""",
                {'key': key},
            )
        except DatabaseUnavailable:
            return []
        return [
            {
                'dish': row['dish'],
                'quantity': float(row['quantity']),
                'base_unit': row['base_unit'],
                'portions': int(row['portions']),
            }
            for row in rows
        ]

    def forget_usage(self, dish: str) -> int:
        """Put a dish's share back - she took it off the menu."""
        try:
            with self.db.connect() as conn:
                cur = conn.cursor()
                cur.execute('DELETE FROM pantry_usage WHERE dish = %(dish)s', {'dish': dish})
                removed = cur.rowcount or 0
                conn.commit()
        except DatabaseUnavailable:
            return 0
        self.reload()
        return removed

    def _package_for(self, key: str, unit: str) -> Package:
        '''Spreadsheet package, overridden by a size learned in conversation.

        A learned size is built directly rather than formatted into a string and
        re-parsed: a float like 400.0 renders with a decimal point, which name
        normalisation turns into a space, which the quantity pattern then reads
        as 0 - and an ingredient whose factor is zero silently drops out of the
        pantry.
        '''
        learned = self.package_sizes.get(key)
        if learned:
            base = UnitConverter.parse(learned['unit'])
            return Package(
                dimension=base.dimension,
                factor=float(learned['quantity']) * base.factor,
                label=f'{learned["quantity"]:g} {learned["unit"]}',
                is_packaged=True,
            )
        try:
            return UnitConverter.parse(unit)
        except UnknownUnitError:
            return UnitConverter.parse('un')

    # ------------------------------------------------------- package sizes

    def _load_package_sizes(self) -> None:
        '''Package weights she told us, which the spreadsheet never records.'''
        if self.db is None:
            return
        try:
            self.package_sizes = {
                row['ingredient_key']: {
                    'quantity': float(row['quantity']),
                    'unit': row['unit'],
                }
                for row in self.db.query('SELECT * FROM package_sizes')
            }
        except DatabaseUnavailable:
            # Still reads correctly; it simply has not learned any weights yet.
            self.package_sizes = {}

    def record_package_size(self, key: str, quantity: float, unit: str) -> None:
        '''Persist how much a loose-piece package actually weighs or holds.'''
        if self.db is not None:
            self.db.execute(
                '''INSERT INTO package_sizes (ingredient_key, quantity, unit)
                        VALUES (%s, %s, %s)
                   ON CONFLICT (ingredient_key) DO UPDATE
                        SET quantity = EXCLUDED.quantity,
                            unit = EXCLUDED.unit,
                            recorded_at = now()''',
                (key, quantity, unit),
            )
        self._load_package_sizes()
        self.reload()

    # ------------------------------------------------------------ querying

    @classmethod
    def _significant_tokens(cls, text: str) -> set[str]:
        return set(text.split()) - cls.STOP_WORDS

    def find(self, name: str) -> PantryItem | None:
        '''Match tolerating accents, case and the sheet parentheticals.

        Returns None rather than a near miss: telling Dona Maria she already
        has an ingredient she does not have is worse than asking.
        '''
        target = UnitConverter.normalise_text(name)
        if target in self.items:
            return self.items[target]

        target_tokens = self._significant_tokens(target)
        if not target_tokens:
            return None

        best, best_score = None, 0.0
        for key, item in self.items.items():
            key_tokens = self._significant_tokens(key)
            if target in key or key in target:
                score = 0.9
            else:
                shared = target_tokens & key_tokens
                if not shared:
                    continue
                score = len(shared) / len(target_tokens | key_tokens)
            if score > best_score:
                best, best_score = item, score
        return best if best_score >= self.MATCH_THRESHOLD else None

    def suggest(self, name: str, limit: int = 5) -> list[str]:
        target_tokens = self._significant_tokens(UnitConverter.normalise_text(name))
        scored = [
            (len(target_tokens & self._significant_tokens(key)), item.name)
            for key, item in self.items.items()
        ]
        scored.sort(reverse=True)
        return [name for score, name in scored[:limit] if score > 0]

    def sorted_items(self) -> list[PantryItem]:
        return sorted(self.items.values(), key=lambda item: item.name)
